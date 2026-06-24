"""Управление складом: остатки, цены, движения, экспорт в Excel."""
from io import BytesIO
from fastapi import APIRouter, Depends, Form, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from app.database import get_db
from app.auth import get_current_user, clean_part_number
from app.models import Inventory, StockTransaction, User
from app.services.cbr_rates import get_latest_rates, convert_to_rub, update_rates

router = APIRouter(prefix="/api/inventory", tags=["inventory"])


def _extras_sum(item: Inventory) -> float:
    """Сумма всех расходов на единицу (в рублях)."""
    return (
        (item.logistics_cost or 0)
        + (item.customs_duty or 0)
        + (item.vat_cost or 0)
        + (item.warehouse_cost or 0)
        + (item.other_costs or 0)
    )


def _full_cost_rub(item: Inventory, rate: float) -> float:
    """Полная себестоимость единицы = закупка*курс + все расходы (₽)."""
    base_purchase_rub = (item.purchase_price or 0) * rate if rate else 0
    return round(base_purchase_rub + _extras_sum(item), 2)


def _recalc_prices(item: Inventory, rate: float):
    """Пересчитать себестоимость и продажные цены.
    Себестоимость = закупка*курс + все расходы.
    Наценка применяется к ПОЛНОЙ себестоимости.
    markup_mode: 'pct' = наценка в %, 'rub' = наценка в рублях, 'manual' = ручная цена.
    """
    cost_rub = _full_cost_rub(item, rate)
    mode = item.markup_mode or "pct"

    if mode == "pct":
        item.price_order = round(cost_rub * (1 + (item.markup_order_pct or 0) / 100), 2)
        item.price_3pl = round(cost_rub * (1 + (item.markup_3pl_pct or 0) / 100), 2)
        item.price_3pl_emex = round(cost_rub * (1 + (item.markup_3pl_emex_pct or 0) / 100), 2)
    elif mode == "rub":
        item.price_order = round(cost_rub + (item.markup_order_rub or 0), 2)
        item.price_3pl = round(cost_rub + (item.markup_3pl_rub or 0), 2)
        item.price_3pl_emex = round(cost_rub + (item.markup_3pl_emex_rub or 0), 2)
    # mode == "manual" — цены не пересчитываются, оставляем как есть

    item.cost_rub = cost_rub
    item.last_rate_used = rate


def _to_float(val, default=0.0) -> float:
    """Безопасное приведение строки/числа из Excel к float."""
    try:
        s = str(val).replace(",", ".").replace(" ", "").strip()
        if not s or s.lower() == "nan":
            return default
        return float(s)
    except Exception:
        return default


@router.get("/")
def list_inventory(
    q: str = "", limit: int = 200,
    db: Session = Depends(get_db), user: User = Depends(get_current_user),
):
    query = db.query(Inventory)
    if q:
        pn = clean_part_number(q)
        query = query.filter(Inventory.part_number_clean.contains(pn))
    items = query.order_by(Inventory.part_number).limit(limit).all()
    rates = get_latest_rates(db)
    return [
        {
            "id": i.id,
            "part_number": i.part_number,
            "brand": i.brand,
            "description": i.description,
            "quantity": i.quantity,
            "purchase_price": i.purchase_price,
            "purchase_currency": i.purchase_currency,
            "purchase_rate": i.purchase_rate,
            "logistics_cost": i.logistics_cost,
            "customs_duty": i.customs_duty,
            "vat_cost": i.vat_cost,
            "warehouse_cost": i.warehouse_cost,
            "other_costs": i.other_costs,
            "extras_total": round(_extras_sum(i), 2),
            "markup_mode": i.markup_mode or "pct",
            "markup_order_pct": i.markup_order_pct,
            "markup_3pl_pct": i.markup_3pl_pct,
            "markup_3pl_emex_pct": i.markup_3pl_emex_pct,
            "markup_order_rub": i.markup_order_rub,
            "markup_3pl_rub": i.markup_3pl_rub,
            "markup_3pl_emex_rub": i.markup_3pl_emex_rub,
            "cost_rub": i.cost_rub,
            "price_order": i.price_order,
            "price_3pl": i.price_3pl,
            "price_3pl_emex": i.price_3pl_emex,
            "margin_order": round(i.price_order - (i.cost_rub or 0), 2) if i.price_order and i.cost_rub else 0,
            "margin_order_pct": round((i.price_order / i.cost_rub - 1) * 100, 1) if i.price_order and i.cost_rub and i.cost_rub > 0 else 0,
            "last_rate_used": i.last_rate_used,
            "updated_at": i.updated_at.isoformat() if i.updated_at else "",
        }
        for i in items
    ]


@router.get("/summary")
def inventory_summary(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    items = db.query(Inventory).all()
    rates = get_latest_rates(db)
    usd_rate = rates.get("USD", 0)

    total_sku = len(items)
    total_qty = sum((i.quantity or 0) for i in items)

    invested_usd = 0.0            # вложено в текущие остатки (только USD-позиции), $
    invested_rub_purchase = 0.0   # стоимость остатков по курсу закупки, ₽
    invested_rub_current = 0.0    # стоимость остатков по текущему курсу, ₽
    cost_total = 0.0              # полная себестоимость остатков, ₽
    rev_order = rev_3pl = rev_emex = 0.0

    for i in items:
        qty = i.quantity or 0
        cur = (i.purchase_currency or "USD").upper()
        cur_rate = rates.get(cur, usd_rate)
        p_rate = i.purchase_rate or cur_rate
        orig = (i.purchase_price or 0) * qty
        invested_rub_purchase += orig * (p_rate or 0)
        invested_rub_current += orig * (cur_rate or 0)
        if cur == "USD":
            invested_usd += orig
        cost_total += (i.cost_rub or 0) * qty
        rev_order += (i.price_order or 0) * qty
        rev_3pl += (i.price_3pl or 0) * qty
        rev_emex += (i.price_3pl_emex or 0) * qty

    # Реализация (фактические продажи)
    sales = db.query(StockTransaction).filter(StockTransaction.tx_type == "sale").all()
    sold_rub = 0.0
    sold_usd_equiv = 0.0
    for s in sales:
        amt = (s.price or 0) * abs(s.quantity or 0)
        sold_rub += amt
        sr = s.sale_rate or usd_rate
        if sr:
            sold_usd_equiv += amt / sr

    fx_diff = round(invested_rub_current - invested_rub_purchase, 2)

    return {
        "total_sku": total_sku,
        "total_quantity": total_qty,
        # Закупка / вложения
        "invested_usd": round(invested_usd, 2),
        "invested_rub_purchase": round(invested_rub_purchase, 2),
        "invested_rub_current": round(invested_rub_current, 2),
        "fx_diff": fx_diff,
        # Полная себестоимость остатков
        "cost_total_rub": round(cost_total, 2),
        # Потенциальная выручка по каналам
        "revenue_order": round(rev_order, 2),
        "revenue_3pl": round(rev_3pl, 2),
        "revenue_emex": round(rev_emex, 2),
        # Потенциальная прибыль по каналам (выручка - себестоимость остатков)
        "profit_order": round(rev_order - cost_total, 2),
        "profit_3pl": round(rev_3pl - cost_total, 2),
        "profit_emex": round(rev_emex - cost_total, 2),
        # Реализация
        "sold_rub": round(sold_rub, 2),
        "sold_usd_equiv": round(sold_usd_equiv, 2),
        "usd_rate": usd_rate,
        # --- совместимость со старым фронтендом ---
        "total_value_usd": round(invested_usd, 2),
        "total_value_rub": round(invested_rub_current, 2),
        "total_revenue_potential": round(rev_order, 2),
    }


@router.post("/update-rates")
async def update_exchange_rates(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """Обновить курсы ЦБ вручную с кнопки."""
    rates = await update_rates(db)
    usd = rates.get("USD", 0)
    eur = rates.get("EUR", 0)
    if not usd:
        return {"error": "Не удалось загрузить курсы ЦБ", "usd": 0, "eur": 0}
    return {"ok": True, "usd": usd, "eur": eur, "total_currencies": len(rates)}


@router.post("/add")
def add_item(
    part_number: str = Form(...),
    brand: str = Form(""),
    description: str = Form(""),
    quantity: int = Form(0),
    purchase_price: float = Form(0),
    purchase_currency: str = Form("USD"),
    logistics_cost: float = Form(0),
    customs_duty: float = Form(0),
    vat_cost: float = Form(0),
    warehouse_cost: float = Form(0),
    other_costs: float = Form(0),
    markup_mode: str = Form("pct"),
    markup_order_pct: float = Form(30),
    markup_3pl_pct: float = Form(35),
    markup_3pl_emex_pct: float = Form(40),
    markup_order_rub: float = Form(0),
    markup_3pl_rub: float = Form(0),
    markup_3pl_emex_rub: float = Form(0),
    price_order_manual: float = Form(0),
    price_3pl_manual: float = Form(0),
    price_3pl_emex_manual: float = Form(0),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    pn_clean = clean_part_number(part_number)
    rates = get_latest_rates(db)
    rate = rates.get(purchase_currency.upper(), rates.get("USD", 0))

    existing = db.query(Inventory).filter(Inventory.part_number_clean == pn_clean).first()
    if existing:
        existing.brand = brand or existing.brand
        existing.description = description or existing.description
        existing.quantity = quantity
        existing.purchase_price = purchase_price
        existing.purchase_currency = purchase_currency
        existing.logistics_cost = logistics_cost
        existing.customs_duty = customs_duty
        existing.vat_cost = vat_cost
        existing.warehouse_cost = warehouse_cost
        existing.other_costs = other_costs
        # курс закупки фиксируем один раз; при повторном добавлении не перетираем
        if not existing.purchase_rate:
            existing.purchase_rate = rate
        existing.markup_mode = markup_mode
        existing.markup_order_pct = markup_order_pct
        existing.markup_3pl_pct = markup_3pl_pct
        existing.markup_3pl_emex_pct = markup_3pl_emex_pct
        existing.markup_order_rub = markup_order_rub
        existing.markup_3pl_rub = markup_3pl_rub
        existing.markup_3pl_emex_rub = markup_3pl_emex_rub
        if markup_mode == "manual":
            existing.price_order = price_order_manual
            existing.price_3pl = price_3pl_manual
            existing.price_3pl_emex = price_3pl_emex_manual
            existing.cost_rub = _full_cost_rub(existing, rate)
            existing.last_rate_used = rate
        else:
            _recalc_prices(existing, rate)
        db.commit()
        return {"ok": True, "action": "updated", "id": existing.id}

    item = Inventory(
        part_number=part_number, part_number_clean=pn_clean,
        brand=brand, description=description, quantity=quantity,
        purchase_price=purchase_price, purchase_currency=purchase_currency,
        purchase_rate=rate,
        logistics_cost=logistics_cost, customs_duty=customs_duty,
        vat_cost=vat_cost, warehouse_cost=warehouse_cost, other_costs=other_costs,
        markup_mode=markup_mode,
        markup_order_pct=markup_order_pct, markup_3pl_pct=markup_3pl_pct,
        markup_3pl_emex_pct=markup_3pl_emex_pct,
        markup_order_rub=markup_order_rub, markup_3pl_rub=markup_3pl_rub,
        markup_3pl_emex_rub=markup_3pl_emex_rub,
    )
    if markup_mode == "manual":
        item.price_order = price_order_manual
        item.price_3pl = price_3pl_manual
        item.price_3pl_emex = price_3pl_emex_manual
        item.cost_rub = _full_cost_rub(item, rate)
        item.last_rate_used = rate
    else:
        _recalc_prices(item, rate)
    db.add(item)
    db.commit()

    if quantity > 0:
        db.add(StockTransaction(
            part_number_clean=pn_clean, tx_type="receipt",
            quantity=quantity, price=purchase_price, sale_rate=rate,
            username=user.username, notes="Первичное поступление",
        ))
        db.commit()

    return {"ok": True, "action": "added", "id": item.id}


@router.post("/upload")
async def upload_inventory(
    file: UploadFile = File(...),
    default_currency: str = Form("USD"),
    default_markup_order: float = Form(30),
    default_markup_3pl: float = Form(35),
    default_markup_3pl_emex: float = Form(40),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    import pandas as pd
    content = await file.read()
    fname = (file.filename or "").lower()
    try:
        if fname.endswith(".csv"):
            df = pd.read_csv(BytesIO(content), dtype=str)
        else:
            df = pd.read_excel(BytesIO(content), dtype=str, engine="openpyxl")
    except Exception as e:
        return {"error": str(e), "added": 0, "updated": 0}

    df = df.dropna(how="all").reset_index(drop=True)
    ncols = len(df.columns)
    rates = get_latest_rates(db)
    rate = rates.get(default_currency.upper(), rates.get("USD", 0))

    added, updated = 0, 0
    for _, row in df.iterrows():
        try:
            pn = str(row.iloc[0]).strip()
            if not pn or pn.lower() == "nan":
                continue
            pn_clean = clean_part_number(pn)
            qty, price, brand, desc = 0, 0.0, "", ""
            logistics = customs = vat = warehouse = 0.0

            # Расширенный формат (9 колонок):
            # артикул, бренд, описание, кол-во, закупка, логистика, пошлина, НДС, склад
            if ncols >= 9:
                brand = str(row.iloc[1]).strip() if str(row.iloc[1]).strip().lower() != 'nan' else ""
                desc = str(row.iloc[2]).strip() if str(row.iloc[2]).strip().lower() != 'nan' else ""
                qty = int(_to_float(row.iloc[3]))
                price = _to_float(row.iloc[4])
                logistics = _to_float(row.iloc[5])
                customs = _to_float(row.iloc[6])
                vat = _to_float(row.iloc[7])
                warehouse = _to_float(row.iloc[8])
            # Базовый формат (5 колонок): артикул, бренд, описание, кол-во, закупка
            elif ncols >= 5:
                brand = str(row.iloc[1]).strip() if str(row.iloc[1]).strip().lower() != 'nan' else ""
                desc = str(row.iloc[2]).strip() if str(row.iloc[2]).strip().lower() != 'nan' else ""
                qty = int(_to_float(row.iloc[3]))
                price = _to_float(row.iloc[4])
            elif ncols >= 3:
                qty = int(_to_float(row.iloc[1]))
                price = _to_float(row.iloc[2])
            elif ncols >= 2:
                qty = int(_to_float(row.iloc[1]))

            existing = db.query(Inventory).filter(Inventory.part_number_clean == pn_clean).first()
            if existing:
                existing.quantity = qty
                if price > 0:
                    existing.purchase_price = price
                if brand:
                    existing.brand = brand
                if desc:
                    existing.description = desc
                if ncols >= 9:
                    existing.logistics_cost = logistics
                    existing.customs_duty = customs
                    existing.vat_cost = vat
                    existing.warehouse_cost = warehouse
                if not existing.purchase_rate:
                    existing.purchase_rate = rate
                _recalc_prices(existing, rate)
                updated += 1
            else:
                item = Inventory(
                    part_number=pn, part_number_clean=pn_clean,
                    brand=brand, description=desc, quantity=qty,
                    purchase_price=price, purchase_currency=default_currency,
                    purchase_rate=rate,
                    logistics_cost=logistics, customs_duty=customs,
                    vat_cost=vat, warehouse_cost=warehouse,
                    markup_order_pct=default_markup_order,
                    markup_3pl_pct=default_markup_3pl,
                    markup_3pl_emex_pct=default_markup_3pl_emex,
                )
                _recalc_prices(item, rate)
                db.add(item)
                added += 1
            db.flush()
        except Exception:
            continue

    db.commit()
    return {"added": added, "updated": updated}


@router.post("/set-expenses")
def set_expenses_bulk(
    logistics_cost: float = Form(0),
    customs_duty: float = Form(0),
    vat_cost: float = Form(0),
    warehouse_cost: float = Form(0),
    other_costs: float = Form(0),
    only_empty: bool = Form(False),
    fields: str = Form("all"),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Массово задать расходы на все позиции.
    fields: 'all' — задать все переданные расходы; либо список через запятую
            из logistics,customs,vat,warehouse,other — задать только их.
    only_empty: если True — заполнять только там, где значение сейчас 0/пусто.
    """
    selected = {f.strip() for f in fields.split(",")} if fields and fields != "all" else None
    mapping = {
        "logistics": ("logistics_cost", logistics_cost),
        "customs": ("customs_duty", customs_duty),
        "vat": ("vat_cost", vat_cost),
        "warehouse": ("warehouse_cost", warehouse_cost),
        "other": ("other_costs", other_costs),
    }
    rates = get_latest_rates(db)
    items = db.query(Inventory).all()
    for item in items:
        for key, (attr, value) in mapping.items():
            if selected is not None and key not in selected:
                continue
            if only_empty and (getattr(item, attr) or 0) > 0:
                continue
            setattr(item, attr, value)
        rate = rates.get((item.purchase_currency or "USD").upper(), rates.get("USD", 0))
        if item.markup_mode == "manual":
            item.cost_rub = _full_cost_rub(item, rate)
            item.last_rate_used = rate
        else:
            _recalc_prices(item, rate)
    db.commit()
    return {"updated": len(items)}


@router.post("/transaction")
def record_transaction(
    part_number: str = Form(...),
    tx_type: str = Form(...),
    quantity: int = Form(...),
    price: float = Form(0),
    notes: str = Form(""),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    pn_clean = clean_part_number(part_number)
    item = db.query(Inventory).filter(Inventory.part_number_clean == pn_clean).first()
    if not item:
        return {"error": "Артикул не найден на складе"}

    if tx_type == "sale":
        if item.quantity < quantity:
            return {"error": f"Недостаточно: есть {item.quantity} шт."}
        item.quantity -= quantity
        qty_change = -quantity
    elif tx_type == "return":
        item.quantity += quantity
        qty_change = quantity
    elif tx_type == "receipt":
        item.quantity += quantity
        qty_change = quantity
    elif tx_type == "adjust":
        qty_change = quantity - item.quantity
        item.quantity = quantity
    else:
        return {"error": "Неизвестный тип операции"}

    # фиксируем текущий курс USD на дату операции (для валютной переоценки продаж)
    rates = get_latest_rates(db)
    cur_rate = rates.get("USD", 0)

    db.add(StockTransaction(
        part_number_clean=pn_clean, tx_type=tx_type,
        quantity=qty_change, price=price, sale_rate=cur_rate,
        notes=notes, username=user.username,
    ))
    db.commit()
    return {"ok": True, "new_quantity": item.quantity}


@router.post("/recalc-prices")
def recalc_all_prices(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    rates = get_latest_rates(db)
    items = db.query(Inventory).all()
    count = 0
    for item in items:
        rate = rates.get((item.purchase_currency or "USD").upper(), rates.get("USD", 0))
        if item.markup_mode == "manual":
            item.cost_rub = _full_cost_rub(item, rate)
            item.last_rate_used = rate
        elif rate > 0:
            _recalc_prices(item, rate)
        count += 1
    db.commit()
    return {"recalculated": count, "usd_rate": rates.get("USD", 0)}


@router.post("/update-markups")
def update_markups(
    markup_mode: str = Form("pct"),
    markup_order_pct: float = Form(30),
    markup_3pl_pct: float = Form(35),
    markup_3pl_emex_pct: float = Form(40),
    markup_order_rub: float = Form(0),
    markup_3pl_rub: float = Form(0),
    markup_3pl_emex_rub: float = Form(0),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    rates = get_latest_rates(db)
    items = db.query(Inventory).all()
    for item in items:
        item.markup_mode = markup_mode
        item.markup_order_pct = markup_order_pct
        item.markup_3pl_pct = markup_3pl_pct
        item.markup_3pl_emex_pct = markup_3pl_emex_pct
        item.markup_order_rub = markup_order_rub
        item.markup_3pl_rub = markup_3pl_rub
        item.markup_3pl_emex_rub = markup_3pl_emex_rub
        rate = rates.get((item.purchase_currency or "USD").upper(), rates.get("USD", 0))
        if rate > 0:
            _recalc_prices(item, rate)
    db.commit()
    return {"updated": len(items)}


@router.post("/delete/{item_id}")
def delete_item(item_id: int, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    item = db.query(Inventory).filter(Inventory.id == item_id).first()
    if item:
        db.delete(item)
        db.commit()
    return {"ok": True}


@router.get("/history/{part_number}")
def get_history(part_number: str, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    pn_clean = clean_part_number(part_number)
    txs = (
        db.query(StockTransaction)
        .filter(StockTransaction.part_number_clean == pn_clean)
        .order_by(StockTransaction.created_at.desc())
        .limit(100).all()
    )
    return [
        {"id": t.id, "type": t.tx_type, "quantity": t.quantity, "price": t.price,
         "sale_rate": t.sale_rate,
         "notes": t.notes, "username": t.username,
         "created_at": t.created_at.isoformat() if t.created_at else ""}
        for t in txs
    ]


@router.get("/export")
def export_excel(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """Внутренний экспорт: полная информация с себестоимостью, расходами и маржой."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    items = db.query(Inventory).order_by(Inventory.part_number).all()
    rates = get_latest_rates(db)
    usd_rate = rates.get("USD", 0)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Остатки"

    headers = [
        "Артикул", "Бренд", "Описание", "Кол-во",
        "Закупка", "Валюта", "Курс закупки", "Курс тек.",
        "Логистика ₽", "Пошлина ₽", "НДС ₽", "Склад ₽", "Прочее ₽",
        "Себестоимость ₽",
        "Заказ под клиента ₽", "Маржа ₽",
        "Склад 3PL ₽", "Маржа ₽",
        "Через EMEX ₽", "Маржа ₽",
    ]
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    thin = Side(style="thin", color="D1D5DB")
    border = Border(bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, item in enumerate(items, 2):
        cost = item.cost_rub or 0
        vals = [
            item.part_number, item.brand, item.description, item.quantity,
            item.purchase_price, item.purchase_currency,
            item.purchase_rate or 0, item.last_rate_used or usd_rate,
            item.logistics_cost or 0, item.customs_duty or 0, item.vat_cost or 0,
            item.warehouse_cost or 0, item.other_costs or 0,
            cost,
            item.price_order, round((item.price_order or 0) - cost, 2),
            item.price_3pl, round((item.price_3pl or 0) - cost, 2),
            item.price_3pl_emex, round((item.price_3pl_emex or 0) - cost, 2),
        ]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=row_idx, column=col, value=v)
            c.border = border

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 32)

    total_row = len(items) + 2
    ws.cell(row=total_row, column=3, value="ИТОГО:").font = Font(bold=True)
    ws.cell(row=total_row, column=4, value=sum((i.quantity or 0) for i in items)).font = Font(bold=True)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=inventory_internal.xlsx"},
    )


@router.get("/export-client")
def export_client_pricelist(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """Клиентский прайс-лист: только артикул, бренд, описание, наличие и 3 цены.
    БЕЗ себестоимости, наценок, курсов и маржи."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from datetime import datetime

    items = db.query(Inventory).order_by(Inventory.part_number).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Прайс-лист"

    # Шапка с названием компании
    title_font = Font(bold=True, size=16, color="1E3A8A")
    sub_font = Font(size=10, color="6B7280")
    ws.merge_cells("A1:G1")
    ws["A1"] = "Dispare Trading"
    ws["A1"].font = title_font
    ws.merge_cells("A2:G2")
    ws["A2"] = "Прайс-лист автозапчастей · " + datetime.now().strftime("%d.%m.%Y")
    ws["A2"].font = sub_font

    headers = [
        "Артикул", "Бренд", "Описание", "Наличие, шт.",
        "Заказ под клиента ₽", "Склад 3PL ₽", "Через EMEX ₽",
    ]
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    thin = Side(style="thin", color="D1D5DB")
    border = Border(bottom=thin)
    HEADER_ROW = 4

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=HEADER_ROW, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for offset, item in enumerate(items, 1):
        row_idx = HEADER_ROW + offset
        avail = item.quantity if (item.quantity or 0) > 0 else "под заказ"
        vals = [
            item.part_number, item.brand, item.description, avail,
            item.price_order or 0, item.price_3pl or 0, item.price_3pl_emex or 0,
        ]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=row_idx, column=col, value=v)
            c.border = border
            if col >= 5:
                c.alignment = Alignment(horizontal="right")

    widths = [18, 16, 40, 14, 20, 18, 18]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    ws.freeze_panes = "A5"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=dispare_pricelist.xlsx"},
    )
