"""Emex-раздел: секретный прайс по ссылке, обработка заказа, дашборд продаж."""
import os
import secrets
import datetime
from io import BytesIO

from fastapi import APIRouter, Request, Depends, UploadFile, File, Body, Form
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse, FileResponse, JSONResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill

from app.database import get_db
from app.auth import get_current_user, clean_part_number
from app.models import User, Inventory, StockTransaction, EmexSetting, MarketPrice, Batch, BatchLot
from app.services import emex_order, batches

router = APIRouter(tags=["emex"])
templates = Jinja2Templates(directory="app/templates")

OUT_DIR = "data/emex_out"
# Emex НЕ удерживает с нас комиссию/страховку/сортировку — он накручивает их
# СВЕРХУ покупателю. Наш расход — только стикеровка (Emex удерживает из платежа).
DEFAULTS = {
    "sticker": "39",        # ₽/шт, с НДС (входящий НДС возмещается)
    "sticker_vat": "1",     # 1 = в стикеровке есть НДС 22% (возмещаемый)
    "markup": "1",          # наша наценка в настройках Emex, %
    "shelf_markup": "46",   # накрутка Emex для анонимного покупателя, % (справочно)
    "start_date": "2026-07-09",   # дата старта продаж (для оборачиваемости)
}


def _user_or_none(request: Request, db: Session):
    try:
        return get_current_user(request, db)
    except Exception:
        return None


def get_setting(db: Session, key: str, default: str = "") -> str:
    row = db.query(EmexSetting).filter(EmexSetting.key == key).first()
    return row.value if row else default


def set_setting(db: Session, key: str, value: str):
    row = db.query(EmexSetting).filter(EmexSetting.key == key).first()
    if row:
        row.value = value
    else:
        db.add(EmexSetting(key=key, value=value))
    db.commit()


def get_or_create_token(db: Session) -> str:
    tok = get_setting(db, "price_token", "")
    if not tok:
        tok = secrets.token_hex(6)
        set_setting(db, "price_token", tok)
    return tok


# ---------------- страница дашборда ----------------
@router.get("/emex", response_class=HTMLResponse)
def emex_page(request: Request, db: Session = Depends(get_db)):
    user = _user_or_none(request, db)
    if not user:
        return RedirectResponse("/login", status_code=302)
    token = get_or_create_token(db)
    return templates.TemplateResponse("emex.html", {
        "request": request, "user": user, "price_token": token,
    })


# ---------------- данные для дашборда (JSON) ----------------
@router.get("/emex/data")
def emex_data(request: Request, db: Session = Depends(get_db),
              user: User = Depends(get_current_user)):
    inv = db.query(Inventory).all()
    by_clean = {i.part_number_clean: i for i in inv}
    market = {}
    for m in db.query(MarketPrice).order_by(MarketPrice.fetched_at.asc()).all():
        market[m.part_number_clean] = m.price

    catalog = []
    for i in inv:
        # сколько приехало всего (по всем партиям) — для «продано % от партии»
        lots = db.query(BatchLot).filter(BatchLot.part_number_clean == i.part_number_clean).all()
        init = sum(l.qty_in for l in lots) or (i.quantity or 0)
        catalog.append({
            "art": i.part_number, "clean": i.part_number_clean,
            "type": i.description or i.brand or "",
            "cost": round(i.cost_rub or 0, 2),
            "price": round(i.price_3pl_emex or 0, 2),
            "stock": i.quantity or 0,
            "init": init,
            "since": i.first_stock_date.isoformat() if i.first_stock_date else None,
            "market": round(market[i.part_number_clean], 2) if market.get(i.part_number_clean) else None,
        })

    sales = []
    q = (db.query(StockTransaction)
           .filter(StockTransaction.tx_type == "sale")
           .order_by(StockTransaction.created_at.asc()))
    for s_ in q.all():
        item = by_clean.get(s_.part_number_clean)
        d_ = s_.op_date or (s_.created_at.date() if s_.created_at else None)
        sales.append({
            "date": (d_.isoformat() if d_ else ""),
            "art": item.part_number if item else s_.part_number_clean,
            "qty": abs(s_.quantity or 0),
            "price": round(s_.price or 0, 2),
            # СНИМОК себестоимости на момент продажи; если пусто (старые записи) — текущая
            "cost": round(s_.cost_at_sale or (item.cost_rub if item else 0) or 0, 2),
            "note": s_.notes or "",
        })

    settings = {k: get_setting(db, k, v) for k, v in DEFAULTS.items()}
    batch_list = [{"id": b.id, "name": b.name,
                   "arrival": b.arrival_date.isoformat() if b.arrival_date else None,
                   "start": b.start_sale_date.isoformat() if b.start_sale_date else None}
                  for b in db.query(Batch).order_by(Batch.id.asc()).all()]
    return {"catalog": catalog, "sales": sales, "settings": settings,
            "batches": batch_list, "token": get_or_create_token(db)}


# ---------------- приём партии ----------------
@router.post("/emex/batch-preview")
async def batch_preview(file: UploadFile = File(...), db: Session = Depends(get_db),
                        user: User = Depends(get_current_user)):
    """Показывает, что произойдёт, НЕ меняя базу."""
    try:
        rows = batches.parse_batch_file(await file.read())
    except Exception as e:
        return JSONResponse({"ok": False, "error": f"Не читается файл: {e}"}, status_code=400)
    prev = []
    for r in rows:
        if r["qty"] <= 0:
            continue
        item = db.query(Inventory).filter(Inventory.part_number_clean == r["clean"]).first()
        if item:
            oq, oc = item.quantity or 0, item.cost_rub or 0
            nq = oq + r["qty"]
            nc = round((oq * oc + r["qty"] * r["cost"]) / nq, 2) if nq else r["cost"]
            prev.append({"art": r["art"], "new": False, "old_qty": oq, "in_qty": r["qty"],
                         "qty": nq, "old_cost": round(oc, 2), "in_cost": round(r["cost"], 2), "cost": nc})
        else:
            prev.append({"art": r["art"], "new": True, "old_qty": 0, "in_qty": r["qty"],
                         "qty": r["qty"], "old_cost": 0, "in_cost": round(r["cost"], 2),
                         "cost": round(r["cost"], 2)})
    return {"ok": True, "rows": prev}


@router.post("/emex/batch-receive")
async def batch_receive(file: UploadFile = File(...), name: str = Form(...),
                        arrival: str = Form(...), start_sale: str = Form(""),
                        db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    try:
        rows = batches.parse_batch_file(await file.read())
        ad = datetime.date.fromisoformat(arrival)
        sd = datetime.date.fromisoformat(start_sale) if start_sale else ad
    except Exception as e:
        return JSONResponse({"ok": False, "error": f"Ошибка: {e}"}, status_code=400)
    batch, report = batches.receive_batch(db, rows, name, ad, sd)
    return {"ok": True, "batch": batch.name, "rows": report,
            "added": sum(1 for r in report if r["new"]),
            "merged": sum(1 for r in report if not r["new"])}


# ---------------- загрузка заказа ----------------
@router.post("/emex/upload")
async def emex_upload(request: Request, file: UploadFile = File(...),
                      db: Session = Depends(get_db),
                      user: User = Depends(get_current_user)):
    content = await file.read()
    try:
        order_date, order_no, lines = emex_order.parse_lqld(content)
    except Exception as e:
        return JSONResponse({"ok": False, "error": f"Не удалось прочитать файл заказа: {e}"}, status_code=400)
    if not lines:
        return JSONResponse({"ok": False, "error": "В файле не найдено строк заказа."}, status_code=400)

    cons = emex_order.consolidate(lines)
    changes, missing = emex_order.apply_order_to_inventory(db, cons, order_date, user.username, order_no)

    os.makedirs(OUT_DIR, exist_ok=True)
    stamp = order_date.strftime("%d%m%y")
    wh_name = f"Отгрузка_ДИСПЭР_{stamp}.xlsx"
    ac_name = f"Задание_бухгалтеру_{stamp}.xlsx"
    with open(os.path.join(OUT_DIR, wh_name), "wb") as f:
        f.write(emex_order.build_warehouse_xlsx(cons, order_date))
    with open(os.path.join(OUT_DIR, ac_name), "wb") as f:
        f.write(emex_order.build_accountant_xlsx(lines))

    return {
        "ok": True,
        "order_date": order_date.strftime("%d.%m.%Y"),
        "req_no": f"1/{stamp}",
        "order_no": order_no,
        "deliver": emex_order.next_business_day(order_date).strftime("%d.%m.%Y"),
        "articles": len(cons),
        "units": sum(v["qty"] for v in cons.values()),
        "changes": [{"art": a, "old": o, "sold": s, "new": n} for a, o, s, n in changes],
        "missing": [{"art": a, "sold": s} for a, s in missing],
        "warehouse_file": wh_name,
        "accountant_file": ac_name,
    }


@router.get("/emex/download")
def emex_download(name: str, request: Request, db: Session = Depends(get_db),
                  user: User = Depends(get_current_user)):
    safe = os.path.basename(name)
    path = os.path.join(OUT_DIR, safe)
    if not os.path.exists(path):
        return JSONResponse({"error": "Файл не найден"}, status_code=404)
    return FileResponse(path, filename=safe,
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ---------------- сохранить цены (полуручной прайс) ----------------
@router.post("/emex/save-price")
def emex_save_price(payload: dict = Body(...), db: Session = Depends(get_db),
                    user: User = Depends(get_current_user)):
    n = 0
    for it in payload.get("items", []):
        clean = clean_part_number(str(it.get("art", "")))
        item = db.query(Inventory).filter(Inventory.part_number_clean == clean).first()
        if item and it.get("price") is not None:
            item.price_3pl_emex = round(float(it["price"]), 2)
            item.markup_mode = "manual"  # чтобы авто-пересчёт не перетёр вашу цену
            n += 1
    db.commit()
    return {"ok": True, "updated": n}


# ---------------- импорт цен из файла прайса ----------------
@router.post("/emex/import-prices")
async def emex_import_prices(file: UploadFile = File(...), apply: str = Form("0"),
                             db: Session = Depends(get_db),
                             user: User = Depends(get_current_user)):
    """Читает файл прайса (№ детали | Марка | Цена детали | Количество)
    и обновляет ТОЛЬКО цены. Остатки не трогает — они ведутся заказами.
    apply=0 — показать разницу, apply=1 — записать."""
    content = await file.read()
    try:
        if content[:2] == b"PK":
            df = pd.read_excel(BytesIO(content), header=0, engine="openpyxl")
        else:
            df = pd.read_excel(BytesIO(content), header=0, engine="xlrd")
    except Exception as e:
        return JSONResponse({"ok": False, "error": f"Не читается файл: {e}"}, status_code=400)

    cols = list(df.columns)
    if len(cols) < 3:
        return JSONResponse({"ok": False, "error": "Ожидаются колонки: № детали, Марка, Цена детали"}, status_code=400)

    rows, missing = [], []
    for _, r in df.iterrows():
        raw = r[cols[0]]
        if pd.isna(raw):
            continue
        art = str(int(raw)) if isinstance(raw, float) and float(raw).is_integer() else str(raw).strip()
        try:
            new_price = round(float(str(r[cols[2]]).replace(",", ".")), 2)
        except Exception:
            continue
        item = db.query(Inventory).filter(Inventory.part_number_clean == clean_part_number(art)).first()
        if not item:
            missing.append(art)
            continue
        old = round(item.price_3pl_emex or 0, 2)
        if abs(old - new_price) < 0.005:
            continue
        rows.append({"art": item.part_number, "old": old, "new": new_price,
                     "diff": round((new_price / old - 1) * 100, 2) if old else 0})
        if apply == "1":
            item.price_3pl_emex = new_price
            item.markup_mode = "manual"
    if apply == "1":
        db.commit()
    return {"ok": True, "applied": apply == "1", "rows": rows, "missing": missing}


# ---------------- сохранить расходы ----------------
@router.post("/emex/save-settings")
def emex_save_settings(payload: dict = Body(...), db: Session = Depends(get_db),
                       user: User = Depends(get_current_user)):
    for k in DEFAULTS:
        if k in payload:
            set_setting(db, k, str(payload[k]))
    return {"ok": True}


@router.get("/emex/orders")
def emex_orders(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """История заказов: группируем продажи по номеру+дате заказа Emex."""
    import re
    inv = {i.part_number_clean: i for i in db.query(Inventory).all()}
    grouped = {}
    q = (db.query(StockTransaction)
           .filter(StockTransaction.tx_type == "sale")
           .order_by(StockTransaction.op_date.asc(), StockTransaction.id.asc()))
    for s_ in q.all():
        d_ = s_.op_date or (s_.created_at.date() if s_.created_at else None)
        m = re.search(r"№\s*(\d+)", s_.notes or "")
        num = m.group(1) if m else ""
        date_str = d_.strftime("%d.%m.%Y") if d_ else "—"
        key = (num, date_str)
        it = inv.get(s_.part_number_clean)
        grouped.setdefault(key, []).append({
            "art": it.part_number if it else s_.part_number_clean,
            "type": (it.description if it else ""),
            "qty": abs(s_.quantity or 0), "price": round(s_.price or 0, 2),
            "cost": round(s_.cost_at_sale or 0, 2),
        })
    out = []
    for (num, date_str), lines in grouped.items():
        title = (f"Заказ №{num} от {date_str}" if num else f"Заказ от {date_str}")
        out.append({"num": num, "date": date_str, "title": title, "lines": lines,
                    "units": sum(l["qty"] for l in lines),
                    "sum": round(sum(l["qty"] * l["price"] for l in lines), 2)})
    return {"orders": out}


@router.get("/emex/orders-export")
def emex_orders_export(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """Экспорт истории заказов в Excel."""
    data = emex_orders(db=db, user=user)["orders"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "История заказов"
    head = ["Заказ №", "Дата", "Артикул", "Наименование", "Кол-во",
            "Цена с НДС ₽/ед", "Сумма с НДС ₽", "Себест. ₽/ед", "Прибыль ₽"]
    for j, h in enumerate(head, 1):
        c = ws.cell(1, j, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="2B4C7E")
    r = 2
    VAT = 1.22
    STICK = float(get_setting(db, "sticker", "39"))
    if get_setting(db, "sticker_vat", "1") == "1":
        STICK = STICK / VAT
    for o in data:
        for ln in o["lines"]:
            profit = (ln["price"] / VAT - ln["cost"] - STICK) * ln["qty"]
            ws.cell(r, 1, o["num"] or "—")
            ws.cell(r, 2, o["date"])
            ws.cell(r, 3, ln["art"])
            ws.cell(r, 4, ln["type"])
            ws.cell(r, 5, ln["qty"])
            ws.cell(r, 6, ln["price"])
            ws.cell(r, 7, round(ln["qty"] * ln["price"], 2))
            ws.cell(r, 8, ln["cost"])
            ws.cell(r, 9, round(profit, 2))
            for col in "FGHI":
                ws[f"{col}{r}"].number_format = "#,##0.00"
            r += 1
    # итог
    ws.cell(r, 4, "ИТОГО").font = Font(bold=True)
    ws.cell(r, 5, f"=SUM(E2:E{r-1})").font = Font(bold=True)
    ws.cell(r, 7, f"=SUM(G2:G{r-1})").font = Font(bold=True)
    ws.cell(r, 9, f"=SUM(I2:I{r-1})").font = Font(bold=True)
    for col in "GI":
        ws[f"{col}{r}"].number_format = "#,##0.00"
    for col, w in zip("ABCDEFGHI", [10, 12, 14, 40, 8, 15, 15, 14, 14]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"История_заказов_Emex_{datetime.date.today().strftime('%d%m%y')}.xlsx"
    from urllib.parse import quote
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(fname)}"})


# ---------------- ПУБЛИЧНЫЙ прайс по секретной ссылке (без логина) ----------------
@router.get("/price/{token}")
def price_feed(token: str, db: Session = Depends(get_db)):
    token = token.removesuffix(".xlsx").removesuffix(".xls").removesuffix(".csv")
    real = get_setting(db, "price_token", "")
    if not real or token != real:
        return JSONResponse({"error": "not found"}, status_code=404)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Прайс"
    for j, h in enumerate(["№ детали", "Марка", "Цена детали", "Количество, шт.", "Товарная группа"], 1):
        ws.cell(1, j, h).font = Font(bold=True)
    r = 2
    for i in db.query(Inventory).all():
        qty = i.quantity or 0
        price = i.price_3pl_emex or 0
        if qty <= 0 or price <= 0:
            continue
        ws.cell(r, 1, i.part_number)
        ws.cell(r, 2, i.brand or "")
        ws.cell(r, 3, round(price, 2))
        ws.cell(r, 4, qty)
        r += 1
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="price.xlsx"'},
    )
