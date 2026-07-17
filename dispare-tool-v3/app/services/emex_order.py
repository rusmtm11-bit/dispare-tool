"""Обработка дневного заказа Emex.

Разбирает файл заказа LQLD, формирует заявку складу (свёрнуто по артикулу)
и построчное задание бухгалтеру, списывает остатки со склада (Inventory)
и пишет продажу в журнал движений (StockTransaction, tx_type='sale').

Ничего не дублирует: остаток и себестоимость берутся из таблицы inventory,
продажи ложатся в stock_transactions — то же, что использует раздел «Склад».
"""
import datetime
from io import BytesIO
from collections import OrderedDict

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy.orm import Session

from app.models import Inventory, StockTransaction, BatchLot
from app.auth import clean_part_number
from app.services.cbr_rates import get_latest_rates

VAT = 0.22

WH = dict(
    warehouse="DTV02", supply="Исходящая поставка", stype="Отгрузка ДИСПЭР",
    owner="ДИСПЭР", partner="EMEX.RU", unit="Короб (BOXWMS)", stock="Свободный запас",
    addr="140091, Московская область, г. Дзержинский, ул. Энергетиков, д. 22, корп. 1.",
)


def next_business_day(d: datetime.date) -> datetime.date:
    nd = d + datetime.timedelta(days=1)
    while nd.weekday() >= 5:
        nd += datetime.timedelta(days=1)
    return nd


def _read_any_excel(content: bytes) -> pd.DataFrame:
    """Читает и старый .xls (OLE2), и .xlsx (zip) из байтов."""
    if content[:2] == b"PK":
        return pd.read_excel(BytesIO(content), header=None, engine="openpyxl")
    return pd.read_excel(BytesIO(content), header=None, engine="xlrd")


def parse_lqld(content: bytes):
    """Возвращает (order_date, lines).
    lines: список [номер, артикул(сырой), артикул(чистый), наименование, кол-во, цена с НДС].
    """
    raw = _read_any_excel(content)
    order_date = None
    for _, r in raw.iterrows():
        for j in range(raw.shape[1]):
            v = r[j]
            if isinstance(v, str) and v.strip().count(".") == 2 and len(v.strip()) == 10:
                try:
                    order_date = datetime.datetime.strptime(v.strip(), "%d.%m.%Y").date()
                except Exception:
                    pass
    if order_date is None:
        order_date = datetime.date.today()

    lines = []
    for _, r in raw.iterrows():
        no, art = r[0], r[3]
        if pd.isna(no) or not isinstance(art, str) or not art.strip() or art == "Total:":
            continue
        try:
            int(str(no).strip())
        except Exception:
            continue
        qty = 0 if pd.isna(r[7]) else int(float(str(r[7]).replace(",", ".")))
        price = None if pd.isna(r[11]) else float(str(r[11]).replace(",", "."))
        name = "" if pd.isna(r[5]) else str(r[5]).strip()
        lines.append([int(str(no).strip()), art.strip(), clean_part_number(art), name, qty, price])
    return order_date, lines


def consolidate(lines):
    cons = OrderedDict()
    for _, raw, clean, name, qty, price in lines:
        if clean in cons:
            cons[clean]["qty"] += qty
        else:
            cons[clean] = {"qty": qty, "raw": raw, "name": name, "price": price}
    return cons


def build_warehouse_xlsx(cons, order_date, seq=1) -> bytes:
    req = f"{seq}/{order_date.strftime('%d%m%y')}"
    deliver = next_business_day(order_date)
    wb = openpyxl.Workbook()
    hs = wb.active
    hs.title = "Заголовок"
    head = ["Номер заявки", "Наименование склада", "Вид поставки", "Тип поставки",
            "Владелец запаса", "Контрагент", "Дата поставки", "Время поставки",
            "Дата доставки", "Адрес доставки (физический)"]
    for j, h in enumerate(head, 1):
        hs.cell(1, j, h).font = Font(bold=True)
    vals = [req, WH["warehouse"], WH["supply"], WH["stype"], WH["owner"], WH["partner"],
            datetime.datetime.combine(order_date, datetime.time()), None,
            datetime.datetime.combine(deliver, datetime.time()), WH["addr"]]
    for j, v in enumerate(vals, 1):
        if v is not None:
            hs.cell(2, j, v)
    ts = wb.create_sheet("Табличная часть")
    thead = ["Номер заявки", "Артикул", "Единица материала", "Количество",
             "Вид запаса", "Владелец запаса (по строке)"]
    for j, h in zip([1, 2, 3, 4, 5, 11], thead):
        ts.cell(1, j, h).font = Font(bold=True)
    r = 2
    for clean, v in cons.items():
        ts.cell(r, 1, req)
        ts.cell(r, 2, v["raw"])
        ts.cell(r, 3, WH["unit"])
        ts.cell(r, 4, v["qty"])
        ts.cell(r, 5, WH["stock"])
        ts.cell(r, 11, WH["owner"])
        r += 1
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_accountant_xlsx(lines) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Для УПД"
    ws["L1"] = VAT
    ws["K1"] = "Ставка НДС →"
    ws["L1"].number_format = "0%"
    ws["K1"].font = Font(bold=True)
    head = ["№ п/п", "Артикул", "Наименование", "Кол-во", "Цена с НДС, ₽/ед",
            "Цена без НДС, ₽/ед", "Стоимость без НДС, ₽", "Ставка НДС",
            "Сумма НДС, ₽", "Стоимость с НДС, ₽"]
    for j, h in enumerate(head, 1):
        c = ws.cell(2, j, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="305496")
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    r = 3
    for no, raw, clean, name, qty, price in lines:
        ws.cell(r, 1, no)
        ws.cell(r, 2, raw)
        ws.cell(r, 3, name)
        ws.cell(r, 4, qty)
        ws.cell(r, 5, price)
        ws.cell(r, 6, f"=E{r}/(1+$L$1)")
        ws.cell(r, 7, f"=F{r}*D{r}")
        ws.cell(r, 8, VAT).number_format = "0%"
        ws.cell(r, 9, f"=G{r}*$L$1")
        ws.cell(r, 10, f"=G{r}+I{r}")
        for col in "EFGIJ":
            ws[f"{col}{r}"].number_format = "#,##0.00"
        r += 1
    ws.cell(r, 3, "ИТОГО").font = Font(bold=True)
    ws.cell(r, 4, f"=SUM(D3:D{r-1})").font = Font(bold=True)
    ws.cell(r, 7, f"=SUM(G3:G{r-1})").font = Font(bold=True)
    ws.cell(r, 9, f"=SUM(I3:I{r-1})").font = Font(bold=True)
    ws.cell(r, 10, f"=SUM(J3:J{r-1})").font = Font(bold=True)
    for col, w in zip("ABCDEFGHIJ", [7, 14, 40, 8, 15, 15, 17, 9, 13, 17]):
        ws.column_dimensions[col].width = w
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def apply_order_to_inventory(db: Session, cons, order_date, username: str):
    """Списывает остатки и пишет продажи.

    Важно: в продажу пишется СНИМОК себестоимости (cost_at_sale) на момент
    сделки. Иначе при приходе новой партии средняя себестоимость изменится
    и прибыль по прошлым продажам пересчитается задним числом.

    Лоты партий (BatchLot) списываются по ФИФО — это нужно, чтобы позже
    можно было включить учёт по ФИФО без потери истории.
    """
    rates = get_latest_rates(db)
    cur_rate = rates.get("USD", 0)
    changes, missing = [], []
    for clean, v in cons.items():
        item = db.query(Inventory).filter(Inventory.part_number_clean == clean).first()
        if not item:
            missing.append((v["raw"], v["qty"]))
            continue
        old_qty = item.quantity or 0
        item.quantity = old_qty - v["qty"]

        # списываем лоты по ФИФО (для будущего ФИФО-учёта)
        left = v["qty"]
        first_lot = 0
        lots = (db.query(BatchLot)
                  .filter(BatchLot.part_number_clean == clean, BatchLot.qty_left > 0)
                  .order_by(BatchLot.received_at.asc(), BatchLot.id.asc()).all())
        for lot in lots:
            if left <= 0:
                break
            take = min(lot.qty_left, left)
            lot.qty_left -= take
            left -= take
            if not first_lot:
                first_lot = lot.batch_id

        db.add(StockTransaction(
            part_number_clean=clean, tx_type="sale",
            quantity=-v["qty"], price=v["price"] or 0,
            cost_at_sale=item.cost_rub or 0,      # снимок: себестоимость на момент продажи
            batch_id=first_lot,
            sale_rate=cur_rate,
            notes=f"Emex заказ {order_date.strftime('%d.%m.%Y')}", username=username,
        ))
        changes.append((v["raw"], old_qty, v["qty"], item.quantity))
    db.commit()
    return changes, missing
